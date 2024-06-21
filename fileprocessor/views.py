import pandas as pd
from django.shortcuts import render
from django.http import HttpResponse
from .forms import UploadFileForm
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def home(request):
    return render(request, 'fileprocessor/home.html')

def charges(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            output = handle_charges(uploaded_file)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Charges_Summary.xlsx'
            return response
    else:
        form = UploadFileForm()
    return render(request, 'fileprocessor/charges.html', {'form': form})

def surgeries(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            output = handle_surgeries(uploaded_file)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Surgeries_Summary.xlsx'
            return response
    else:
        form = UploadFileForm()
    return render(request, 'fileprocessor/surgeries.html', {'form': form})

def handle_charges(file):
    # Read main sheet, calculate aggregate sum
    if(file.name.endswith('.xls') | file.name.endswith('.xlsx')):
        df = pd.read_excel(file)

    df['SUB GL'] = 'SUB GL ' + df['SUB GL'].astype(str)
    aggregate_sum = df.groupby('SUB GL')['ExtCost'].sum().reset_index()
    total_sum = aggregate_sum['ExtCost'].sum().round(2)
    aggregate_sum = pd.concat([pd.DataFrame({'SUB GL': ['Total Ext Cost'], 'ExtCost': [total_sum]}), aggregate_sum])
    aggregate_sum['Source'] = 'Aggregate'

    # Sheets to be processed (assumes sheet names will remain same over time)
    sheets = ['CIV', 'RESOLUTE', 'SLB', 'MTB', 'NEB', 'NCB', 'BMC']
    all_data = [aggregate_sum]

    # Process each sheet and calculate sum per subgroup
    for sheet in sheets:
        df_sheet = pd.read_excel(file, sheet_name=sheet)
        df_sheet['SUB GL'] = 'SUB GL ' + df_sheet['SUB GL'].astype(str)
        sum_sheet = df_sheet.groupby('SUB GL')['ExtCost'].sum().reset_index()
        total_sum = sum_sheet['ExtCost'].sum().round(2)
        sum_sheet = pd.concat([pd.DataFrame({'SUB GL': ['Total Ext Cost'], 'ExtCost': [total_sum]}), sum_sheet])
        sum_sheet['Facility'] = sheet
        all_data.append(sum_sheet)

    # Convert 'ExtCost' columns to accounting format
    for df in all_data:
        columns_to_format = [col for col in df.columns if 'ExtCost' in col]
        for col in columns_to_format:
            df[col] = df[col].apply(lambda x: f'${x:,.2f}' if pd.notnull(x) else x)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Styling definitions
    header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    cell_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    font = Font(bold=True, color="000000")
    alignment_header = Alignment(horizontal="center")
    alignment_accounting = Alignment(horizontal="right")

    # Write each dataframe to Excel in different columns
    start_col = 1
    for df in all_data:
        # Write each header and style header
        for col_num, column_title in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.fill = header_fill
            cell.font = font
            cell.alignment = alignment_header

        for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for col_num, cell_value in enumerate(row, start=start_col):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                if df.columns[col_num - start_col] in ['SUB GL', 'ExtCost', 'Facility']:
                    cell.fill = cell_fill
                    if df.columns[col_num - start_col] in columns_to_format:
                        cell.alignment = alignment_accounting

        # Auto adjust column width to ensure readability
        for col_num, column_title in enumerate(df.columns, start=start_col):
            max_length = max(
                len(str(column_title)),
                *(len(str(cell_value)) for cell_value in df[column_title])
            )
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_length + 2

        # Add break in columns
        start_col += len(df.columns) + 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def handle_surgeries(file):
    if file.name.endswith('.csv'):
        df = pd.read_csv(file)
    elif file.name.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(file)
    else:
        raise ValueError("Unsupported file format. Only CSV and XLS/XLSX are supported.")

    # Removes duplicate event IDs in column EVENT_ID
    df_unique = df.drop_duplicates(subset=['EVENT_ID'])

    # Array of surgeon names
    # Might need to change depending on how surgeon names are profiled
    surgeon_names = [
        "BURGESS MD, MEGAN", "CYR MD, STEVEN", "DEBERARDINO MD, THOMAS", 
        "FERGUSON MD, EARL", "GARCIA MD, FRANCISCO", "KAISER MD, BRYAN", 
        "KREINES DO, ALEXANDER", "LYNCH MD, JAMIE", "NILSSON MD, JOEL", 
        "SWANN MD, MATTHEW", "VIROSLAV MD, SERGIO"
    ]

    # Surgeons that require 4 bags per case
    multiply_by_4 = ["DEBERARDINO", "KREINES", "NILSSON", "LYNCH"]

    surgeon_count = {}

    for surgeon_name in surgeon_names:
        last_name = surgeon_name.split(' ')[0] # Get last name of surgeon
        count = df_unique[df_unique['SURGEON'].str.contains(surgeon_name, na=False)].shape[0] # Count number of occurrences of each surgeon
        total_bags = count * 4 if last_name in multiply_by_4 else count # Calculate total bags for each surgeon
        surgeon_count[last_name] = (count, total_bags)

    # Array of breast surgeons
    breast_surgeons = [
        "CORNEJO MD", "GASSMAN MD, ANDREW", "OCHOA MD, OSCAR", "WHIPPLE MD, LAUREN"
    ]

    breast_count = {surgeon.split(" ")[0]: 0 for surgeon in breast_surgeons}

    # Filter the DataFrame for breast surgeons
    breast_df = df[df['SURGEON'].apply(lambda x: any(name in x for name in breast_surgeons))]

    # Dictionary to track seen EVENT_IDs
    seen_id = set()

    # Iterate over the filtered DataFrame
    for index, row in breast_df.iterrows():
        event_id = row['EVENT_ID']
        surgeon = row['SURGEON']
        proc_text = row['PROC_TEXT']
        
        # Check if EVENT_ID has already been seen
        if event_id in seen_id:
            continue
        
        # Check if "BREAST" is in the PROC_TEXT
        if "BREAST" in proc_text.upper():
            last_name = surgeon.split(" ")[0]
            breast_count[last_name] += 1
            seen_id.add(event_id)
    
    for index, row in breast_df.iterrows():
        print(index, row)

    # Count CABG and TAVR cases for the week
    CABG_count = df_unique['PROC_TEXT'].str.contains('CABG', case=False, na=False).sum()
    TAVR_count = df_unique['PROC_TEXT'].str.contains('TAVR', case=False, na=False).sum()

    # Write results to Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Surgery Summary"

    # Styling definitions
    header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    font = Font(bold=True, color="000000")

    # Write headers
    ws.append(['Surgeon Last Name', 'Count', 'Total Bags', 'Bag Name'])
    header_row = ws[1]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = font

    # Write data
    for last_name, count in surgeon_count.items():
        if(last_name == "BURGESS"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, 'Tumescent Solution – NS 1000ml + TXA 1gm + Lidocaine 1% 50ml + Epi 1ml. Profile but wait for request to mix due to low stability.'])
            continue
        if(last_name == "CYR"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, 'Heparin 30,000 unit in 1000ml NS for Cell Saver.'])
            continue
        if(last_name == "DEBERARDINO"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1mg Epinephrine in NS 3000ml irrigation'])
            continue
        if(last_name == "FERGUSON"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '500,000 unit Polymyxin, Ancef 1gram, Gent 80mg in NS 1 liter IRRIGATION'])
            continue
        if(last_name == "GARCIA"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1,000,000 units Polymyxin in NS 3000ml irrigation bag'])
            continue
        if(last_name == "KAISER"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1,000,000 units Polymyxin in NS 3000ml irrigation bag'])
            continue
        if(last_name == "KREINES"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1mg Epinephrine in NS 3000ml irrigation bag'])
            continue
        if(last_name == "LYNCH"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1mg EPI in NS 3000ml irrigation bag x 4 bags each case (NOT SPINE)'])
            continue
        if(last_name == "NILSSON"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1mg Epinephrine in NS 3000ml irrigation bag'])
            continue
        if(last_name == "SWANN"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, 'Heparin 30,000 unit in 1000ml NS for Cell Saver'])
            continue
        if(last_name == "VIROSLAV"):
            original_count, total_bags = count
            ws.append([last_name, original_count, total_bags, '1,000,000 units Polymyxin in NS 3000ml irrigation bag'])
            continue
        
        original_count, total_bags = count
        ws.append([last_name, original_count, total_bags])

    ws.append([])

    ws.append(["Surgeon Last Name", "Breast Cases Count", "Total Bags", "Bag Name"])

    header_row2 = ws[14]
    for cell in header_row2:
        cell.fill = header_fill
        cell.font = font

    for surgeon, count in breast_count.items():
        total_bags = count
        if(surgeon == "CORNEJO"):
            ws.append([surgeon, count, total_bags, "Tumescent Solution – LR 1000ml + Lidocaine 1% 50ml + Epi 1ml: Profile but wait for request to mix due to low stability."])
        if(surgeon == "GASSMAN"):
            ws.append([surgeon, count, total_bags, "PRMA solutions: Papaverine 120mg in NS 100ml and Heparin 20,000 unit in NS 500mL."])
        if(surgeon == "OCHOA"):
            ws.append([surgeon, count, total_bags, "PRMA solutions: Papaverine 120mg in NS 100ml and Heparin 20,000 unit in NS 500mL."])
        if(surgeon == "WHIPPLE"):
            ws.append([surgeon, count, total_bags, "PRMA solutions: Papaverine 120mg in NS 100ml and Heparin 20,000 unit in NS 500mL."])

    ws.append([]) # Creates empty row

    ws.append(["CABG Cases This Week: ", CABG_count]) # CABG output
    ws.append(["TAVR Cases This Week: ", TAVR_count]) # TAVR output

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
